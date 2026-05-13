import streamlit as st
import pandas as pd
import numpy as np
from fpdf import FPDF
from datetime import date, datetime
import tempfile
import requests
import os
from io import BytesIO
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

# --- CONFIGURACIÓN DE COLORES Y RUTAS ---
COLOR_VERDE = "#008f39"   
COLOR_GEMINI_DARK = "#1e1f20" 
COLOR_TEXTO = "#e3e3e3"

URL_LOGO_WEB = "https://raw.githubusercontent.com/Osfelp/IMAGENES-ELYON-/main/logo-elyon-GRIS.png"
URL_LOGO_PDF = "https://raw.githubusercontent.com/Osfelp/IMAGENES-ELYON-/main/logo-elyon.png"
URL_WATERMARK_PDF = "https://raw.githubusercontent.com/Osfelp/IMAGENES-ELYON-/main/logo-elyon-GRIS2%2B.png"

st.set_page_config(page_title="Formato de Cotización | Elyon", page_icon="🎬", layout="wide")

# --- DESCARGA RÁPIDA DE IMÁGENES ---
@st.cache_data(show_spinner=False)
def obtener_imagen_bytes(url):
    try:
        req = requests.get(url)
        if req.status_code == 200: return req.content
    except: pass
    return None

# --- DISEÑO (CSS PREMIUM UI/UX) ---
st.markdown(f"""
    <style>
    .stApp {{ background-color: {COLOR_GEMINI_DARK}; color: {COLOR_TEXTO}; }}
    h1, h2, h3, h4, h5, label, .stMetric {{ color: {COLOR_VERDE} !important; font-family: 'Lexend', sans-serif; }}
    
    .stTextInput>div>div>input, .stNumberInput>div>div>input, .stDateInput>div>div>input {{
        background-color: #161718 !important; color: white !important;
        border: 1px solid #3a3b3c !important; border-radius: 8px !important; padding: 8px 12px !important;
    }}
    [data-testid="stDataEditor"] {{ background-color: #161718; border: 1px solid #3a3b3c; border-radius: 10px; }}
    
    [data-testid="metric-container"] {{
        background: linear-gradient(145deg, #1e1f20 0%, #111213 100%);
        border: 1px solid #3a3b3c; padding: 20px 15px !important; border-radius: 12px !important;
        box-shadow: 0px 6px 15px rgba(0,0,0,0.3); text-align: center;
    }}
    [data-testid="metric-container"] label {{ justify-content: center; font-size: 0.95rem !important; opacity: 0.8; margin-bottom: 5px; }}
    [data-testid="metric-container"] div {{ justify-content: center; font-size: 2.1rem !important; color: white !important; font-weight: bold; }}

    div[data-testid="stHorizontalBlock"]:last-of-type [data-testid="column"]:nth-child(1) button {{
        background: linear-gradient(135deg, #008f39 0%, #00501f 100%) !important; 
        color: white !important; height: 3.8em !important; border-radius: 8px !important; border: none !important; font-weight: bold !important; box-shadow: 0 4px 10px rgba(0,0,0,0.4) !important;
    }}
    div[data-testid="stHorizontalBlock"]:last-of-type [data-testid="column"]:nth-child(2) button {{
        background: linear-gradient(135deg, #d35400 0%, #903a00 100%) !important; 
        color: white !important; height: 3.8em !important; border-radius: 8px !important; border: none !important; font-weight: bold !important; box-shadow: 0 4px 10px rgba(0,0,0,0.4) !important;
    }}
    div[data-testid="stHorizontalBlock"]:last-of-type [data-testid="column"]:nth-child(3) button {{
        background: linear-gradient(135deg, #2980b9 0%, #154360 100%) !important; 
        color: white !important; height: 3.8em !important; border-radius: 8px !important; border: none !important; font-weight: bold !important; box-shadow: 0 4px 10px rgba(0,0,0,0.4) !important;
    }}
    div[data-testid="stHorizontalBlock"]:last-of-type button:hover {{ transform: translateY(-2px); transition: 0.2s; }}

    .sidebar-card {{ background-color: #161718; border: 1px solid #3a3b3c; padding: 18px; border-radius: 12px; margin-bottom: 10px; box-shadow: 0px 4px 12px rgba(0,0,0,0.4); }}
    [data-testid="stSidebar"] [data-testid="stDownloadButton"] button {{ background: linear-gradient(135deg, #2980b9 0%, #154360 100%) !important; color: white !important; border: none !important; font-weight: bold !important; height: 3em !important; }}
    
    .btn-edit button {{ background-color: #2c3e50 !important; color: white !important; height: 36px !important; padding: 0px 15px !important; border-radius: 8px !important; border: none !important; margin-top: 13px !important; }}
    .btn-edit button:hover {{ background-color: #34495e !important; transform: scale(1.05); }}
    
    .btn-delete button {{ background-color: #7b241c !important; color: white !important; height: 36px !important; padding: 0px 15px !important; border-radius: 8px !important; border: none !important; margin-top: 13px !important; }}
    .btn-delete button:hover {{ background-color: #922b21 !important; transform: scale(1.05); }}
    
    .group-header {{ background: linear-gradient(90deg, #161718 0%, #1e1f20 100%); padding: 12px 20px; border-radius: 8px; border: 1px solid #3a3b3c; border-left: 5px solid {COLOR_VERDE}; margin-top: 20px; color: white; font-weight: 700; font-size: 1.1em; }}
    .preview-box {{ background-color: #161718; padding: 25px; border-radius: 12px; border: 1px solid #3a3b3c; margin-top: 15px; box-shadow: 0px 4px 10px rgba(0,0,0,0.2); }}
    </style>
""", unsafe_allow_html=True)

# --- FUNCIONES MATEMÁTICAS ---
def formato_moneda(valor):
    try: return f"$ {int(round(float(valor))):,}".replace(",", ".")
    except: return "$ 0"

def redondear_al_mil(valor):
    try:
        if pd.isna(valor): return 0
        v = float(valor); return 0 if v <= 0 else int(math.ceil(v / 1000.0) * 1000)
    except: return 0

# --- CLASE PDF CON FOOTER ELEGANTE ---
class PDF(FPDF):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.set_auto_page_break(auto=True, margin=20) 
    def footer(self):
        self.set_y(-18)
        self.set_draw_color(200, 200, 200)
        self.line(15, self.get_y(), 200, self.get_y()) # Línea separadora
        self.ln(2)
        self.set_font('Arial', 'B', 8)
        self.set_text_color(0, 143, 57) # Verde Elyon
        self.cell(0, 4, "ELYON PRODUCCIONES S.A.S.", 0, 1, 'C')
        self.set_font('Arial', '', 8)
        self.set_text_color(100, 100, 100)
        self.cell(0, 4, "Cl 136 # 45-35 Bogotá, Colombia  |  www.elyonproducciones.com", 0, 0, 'C')

DEFAULT_VALS = {"Descripción": "", "Días": 1.0, "Cantidad": 1.0, "Costo Unitario": 0.0, "Incremento (%)": 40.0}

# --- SISTEMA DE RECARGA ---
if "uploaded_filename" not in st.session_state:
    st.session_state.uploaded_filename = ""

st.sidebar.markdown("""
<div class="sidebar-card">
    <h3 style='color: #008f39; margin-bottom: 8px; font-size: 1.1em;'>🔄 RECUPERAR DATOS</h3>
    <p style='color: #bbb; font-size: 0.85em; line-height: 1.4; margin-bottom: 0;'>Sube un Excel <b>(.xlsx)</b> descargado desde esta plataforma para seguir editándolo sin empezar de cero.</p>
</div>
""", unsafe_allow_html=True)
uploaded_file = st.sidebar.file_uploader("", type=["xlsx"], label_visibility="collapsed")

if uploaded_file is not None:
    if uploaded_file.name != st.session_state.uploaded_filename:
        try:
            df_info = pd.read_excel(uploaded_file, sheet_name="ELYON_BACKUP", header=None, nrows=18)
            info_dict = dict(zip(df_info[0], df_info[1]))
            st.session_state['val_cot'] = str(info_dict.get('cotizacion_num', 'ELY-'))
            st.session_state['val_ciu'] = str(info_dict.get('ciudad', ''))
            try: st.session_state['val_fec'] = datetime.strptime(str(info_dict.get('fecha_emision')), '%Y-%m-%d').date()
            except: pass
            st.session_state['val_val'] = int(float(info_dict.get('validez_dias', 15)))
            st.session_state['val_sinf'] = str(info_dict.get('sin_fee')) == 'True'
            st.session_state['val_fee'] = float(info_dict.get('fee_porcentaje_input', 10.0))
            st.session_state['val_iva'] = float(info_dict.get('iva_porcentaje', 19.0))
            st.session_state['val_emp'] = str(info_dict.get('solicitante_empresa', ''))
            st.session_state['val_con'] = str(info_dict.get('solicitante_contacto', ''))
            st.session_state['val_ema'] = str(info_dict.get('solicitante_email', ''))
            st.session_state['val_tel'] = str(info_dict.get('solicitante_telefono', ''))
            st.session_state['val_ref'] = str(info_dict.get('evento_referencia', ''))
            try: st.session_state['val_fev'] = datetime.strptime(str(info_dict.get('evento_fecha')), '%Y-%m-%d').date()
            except: pass
            st.session_state['val_lug'] = str(info_dict.get('evento_lugar', ''))
            st.session_state['val_asi'] = str(info_dict.get('evento_asistentes', ''))

            for k in ['val_ciu', 'val_emp', 'val_con', 'val_ema', 'val_tel', 'val_ref', 'val_lug', 'val_asi']:
                if str(st.session_state.get(k)) == 'nan': st.session_state[k] = ''

            df_raw = pd.read_excel(uploaded_file, sheet_name="ELYON_BACKUP", skiprows=19)
            st.session_state.cat_order = []
            st.session_state.categorias = {}
            for cat in df_raw['Categoría'].dropna().unique():
                st.session_state.cat_order.append(cat)
                st.session_state.categorias[cat] = df_raw[df_raw['Categoría'] == cat][["Descripción", "Días", "Cantidad", "Costo Unitario", "Incremento (%)"]].copy()
            
            st.session_state.uploaded_filename = uploaded_file.name
            st.sidebar.success("✅ Datos recuperados con éxito!")
            st.rerun()
        except:
            st.sidebar.error("Archivo no compatible.")

# --- ENCABEZADO ---
col_l, col_t = st.columns([1, 2.5])
with col_l:
    try: st.image(URL_LOGO_WEB, width=320)
    except: st.write("ELYON")
with col_t:
    st.markdown("<h1 style='margin-top: 10px; font-size: 3em;'>Formato de cotización</h1>", unsafe_allow_html=True)
    st.write("Plataforma de Gestión Comercial | **ELYON PRODUCCIONES**")

st.markdown("<hr>", unsafe_allow_html=True)

# --- I. DATOS GENERALES ---
st.markdown("### I. DATOS GENERALES")
c1, c2, c3, c4, c5, c6, c7 = st.columns([1.5, 1.5, 1.5, 1, 0.8, 0.8, 1])
cotizacion_num = c1.text_input("COTIZACIÓN NO.", value=st.session_state.get("val_cot", "ELY-"), key="val_cot")
ciudad = c2.text_input("CIUDAD", value=st.session_state.get("val_ciu", ""), key="val_ciu")
fecha_emision = c3.date_input("FECHA DE EMISIÓN", value=st.session_state.get("val_fec", date.today()), key="val_fec")
validez_dias = c4.number_input("VALIDEZ (DÍAS)", min_value=1, value=st.session_state.get("val_val", 15), key="val_val")
with c7:
    st.markdown("<div style='margin-top: 32px;'></div>", unsafe_allow_html=True)
    sin_fee = st.checkbox("Sin Fee", value=st.session_state.get("val_sinf", False), key="val_sinf")
fee_porcentaje_input = c5.number_input("FEE %", min_value=0.0, value=float(st.session_state.get("val_fee", 10.0)), disabled=sin_fee, key="val_fee")
fee_pct_final = 0 if sin_fee else fee_porcentaje_input
iva_porcentaje = c6.number_input("IVA %", min_value=0.0, value=float(st.session_state.get("val_iva", 19.0)), key="val_iva") 

c8, c9, c10, c11 = st.columns(4)
solicitante_empresa = c8.text_input("EMPRESA", value=st.session_state.get("val_emp", ""), key="val_emp")
solicitante_contacto = c9.text_input("CONTACTO", value=st.session_state.get("val_con", ""), key="val_con")
solicitante_email = c10.text_input("EMAIL", placeholder="ejemplo@correo.com", value=st.session_state.get("val_ema", ""), key="val_ema")
solicitante_telefono = c11.text_input("TELÉFONO", placeholder="Ej: 3001234567", value=st.session_state.get("val_tel", ""), key="val_tel")

c12, c13, c14, c15 = st.columns(4)
evento_referencia = c12.text_input("REFERENCIA EVENTO", value=st.session_state.get("val_ref", ""), key="val_ref")
evento_fecha = c13.date_input("FECHA EVENTO", value=st.session_state.get("val_fev", date.today()), key="val_fev") 
evento_lugar = c14.text_input("LUGAR EVENTO", value=st.session_state.get("val_lug", ""), key="val_lug")
evento_asistentes = c15.text_input("No. ASISTENTES", value=st.session_state.get("val_asi", ""), key="val_asi")

# --- II. GESTIÓN POR CATEGORÍAS ---
st.markdown("<hr>", unsafe_allow_html=True)
st.markdown("### II. DETALLE DEL SERVICIO")

if 'cat_order' not in st.session_state: st.session_state.cat_order = ["General"]
if 'categorias' not in st.session_state: st.session_state.categorias = {"General": pd.DataFrame([DEFAULT_VALS])}
if 'edited_categorias' not in st.session_state: st.session_state.edited_categorias = {}

def agregar_nueva_seccion():
    n = st.session_state.nueva_cat_input.strip()
    if n and n not in st.session_state.cat_order:
        st.session_state.cat_order.append(n)
        st.session_state.categorias[n] = pd.DataFrame([DEFAULT_VALS])
    st.session_state.nueva_cat_input = ""

def guardar_edicion(old_name, idx):
    new_name = st.session_state[f"edit_input_{old_name}"].strip()
    if new_name and new_name != old_name and new_name not in st.session_state.categorias:
        st.session_state.categorias[new_name] = st.session_state.categorias.pop(old_name)
        st.session_state.cat_order[idx] = new_name
    st.session_state[f"edit_mode_{old_name}"] = False

col_nueva1, col_nueva2 = st.columns([10, 1])
col_nueva1.text_input("Añadir nueva sección...", label_visibility="collapsed", placeholder="Escribe un título y presiona Enter...", key="nueva_cat_input", on_change=agregar_nueva_seccion)
st.markdown("<br>", unsafe_allow_html=True)

df_global = pd.DataFrame()

config_columnas = {
    "Descripción": st.column_config.TextColumn("Descripción", width="large"),
    "Días": st.column_config.NumberColumn("Días", width="small", min_value=0.0, step=1.0, default=1.0),
    "Cantidad": st.column_config.NumberColumn("Cant.", width="small", min_value=0.0, step=1.0, default=1.0),
    "Costo Unitario": st.column_config.NumberColumn("Costo Unitario ($)", width="medium", step=1.0, default=0.0), 
    "Incremento (%)": st.column_config.NumberColumn("Incr. (%)", width="small", step=1.0, default=40.0)
}

for i, cat_name in enumerate(list(st.session_state.cat_order)):
    col_tit, col_edit, col_del = st.columns([8, 0.7, 0.7])
    edit_key = f"edit_mode_{cat_name}"
    
    if st.session_state.get(edit_key, False):
        with col_tit: 
            st.markdown("<div style='margin-top: 10px;'></div>", unsafe_allow_html=True)
            st.text_input("Renombrar sección:", value=cat_name, key=f"edit_input_{cat_name}", on_change=guardar_edicion, args=(cat_name, i), label_visibility="collapsed")
    else:
        with col_tit: st.markdown(f"<h4 style='margin-top: 15px; margin-bottom: 0px; color: {COLOR_VERDE};'>📁 {cat_name.upper()}</h4>", unsafe_allow_html=True)
            
    with col_edit:
        if not st.session_state.get(edit_key, False):
            st.markdown('<div class="btn-edit">', unsafe_allow_html=True)
            if st.button("✏️", key=f"btn_edit_{cat_name}"):
                st.session_state[edit_key] = True; st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

    with col_del:
        if cat_name != "General":
            st.markdown('<div class="btn-delete">', unsafe_allow_html=True)
            if st.button("🗑️", key=f"del_{cat_name}"):
                st.session_state.cat_order.pop(i)
                st.session_state.categorias.pop(cat_name, None); st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
    
    df_base = st.session_state.categorias.get(cat_name, pd.DataFrame([DEFAULT_VALS]))
    edited_df = st.data_editor(df_base, num_rows="dynamic", use_container_width=True, hide_index=True, key=f"editor_{cat_name}", column_config=config_columnas)
    st.session_state.edited_categorias[cat_name] = edited_df.copy()
    
    df_calc = edited_df.copy()
    df_calc["Días"] = pd.to_numeric(df_calc["Días"], errors='coerce').fillna(1.0)
    df_calc["Cantidad"] = pd.to_numeric(df_calc["Cantidad"], errors='coerce').fillna(1.0)
    df_calc["Costo Unitario"] = pd.to_numeric(df_calc["Costo Unitario"], errors='coerce').fillna(0.0)
    df_calc["Incremento (%)"] = pd.to_numeric(df_calc["Incremento (%)"], errors='coerce').fillna(40.0)
    
    precio_base = df_calc["Costo Unitario"] / (1 - (df_calc["Incremento (%)"] / 100))
    df_calc["Precio Unitario"] = (np.ceil(precio_base / 1000.0) * 1000).fillna(0).astype(int)
    subtotal_base = df_calc["Días"] * df_calc["Cantidad"] * df_calc["Precio Unitario"]
    df_calc["Subtotal Item"] = (np.ceil(subtotal_base / 1000.0) * 1000).fillna(0).astype(int)
    df_calc["Categoría"] = cat_name 
    df_global = pd.concat([df_global, df_calc], ignore_index=True)

# --- CÁLCULOS GLOBALES ---
df_validos = df_global[df_global["Descripción"].str.strip() != ""] if not df_global.empty else pd.DataFrame()
subtotal_neto = redondear_al_mil(df_validos["Subtotal Item"].sum()) if not df_validos.empty else 0
fee_total = redondear_al_mil(subtotal_neto * (fee_pct_final / 100.0))
iva_total = redondear_al_mil((subtotal_neto + fee_total) * (iva_porcentaje / 100.0))
total_general = redondear_al_mil(subtotal_neto + fee_total + iva_total)

# --- III. PREVISUALIZACIÓN ---
st.markdown("<hr>", unsafe_allow_html=True)
st.markdown("### 📄 III. PREVISUALIZACIÓN DE COTIZACIÓN")
st.markdown('<div class="preview-box">', unsafe_allow_html=True)
if not df_validos.empty:
    for cat_name in st.session_state.cat_order:
        subset = df_validos[df_validos["Categoría"] == cat_name]
        if not subset.empty:
            st.markdown(f'<div class="group-header">{cat_name.upper()}</div>', unsafe_allow_html=True)
            subset_format = subset[["Descripción", "Días", "Cantidad", "Precio Unitario", "Subtotal Item"]].copy()
            subset_format["Precio Unitario"] = subset_format["Precio Unitario"].apply(formato_moneda)
            subset_format["Subtotal Item"] = subset_format["Subtotal Item"].apply(formato_moneda)
            st.dataframe(subset_format, use_container_width=True, hide_index=True)
st.markdown("<br><br>", unsafe_allow_html=True)

res1, res2, res3, res4 = st.columns(4)
res1.metric("SUBTOTAL NETO", formato_moneda(subtotal_neto))
res2.metric(f"FEE {fee_pct_final}%", formato_moneda(fee_total))
res3.metric(f"IVA {iva_porcentaje}%", formato_moneda(iva_total))
res4.metric("TOTAL GENERAL", formato_moneda(total_general))
st.markdown('</div>', unsafe_allow_html=True)

# --- GENERADOR EXCEL CON TOTALES RESALTADOS ---
def generar_excel_estilizado(df_data, es_control=False):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    currency_format = '$ #,##0'
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="008f39" if not es_control else "2980b9", end_color="008f39" if not es_control else "2980b9", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    curr_row = 1
    ws.cell(curr_row, 1, "CUADRO DE CONTROL INTERNO" if es_control else "COTIZACIÓN COMERCIAL").font = Font(bold=True, size=14); curr_row += 2
    
    # RESTAURAMOS LAS VARIABLES DE CÁLCULO (Evita el Error)
    total_costo_general = 0
    total_venta_general = 0
    total_ganancia_general = 0
    
    if not df_data.empty:
        for cat in st.session_state.cat_order:
            if cat not in df_data["Categoría"].values: continue
            ws.cell(curr_row, 1, f"SECCIÓN: {cat.upper()}").font = Font(bold=True); curr_row += 1
            headers = ["DESCRIPCIÓN", "DIAS", "CANT.", "COSTO UNIT.", "COSTO TOTAL", "PRECIO VENTA UNIT.", "PRECIO VENTA", "GANANCIA", "% GANANCIA"] if es_control else ["DESCRIPCIÓN", "DIAS", "CANT.", "PRECIO UNITARIO", "SUBTOTAL"]
            for col_idx, text in enumerate(headers, 1):
                cell = ws.cell(curr_row, col_idx, text); cell.font, cell.fill, cell.border, cell.alignment = header_font, header_fill, border, Alignment(horizontal="center")
            curr_row += 1
            
            subset = df_data[df_data["Categoría"] == cat]
            sub_costo, sub_venta, sub_gan = 0, 0, 0 # Restaurado el reinicio de subtotales
            
            for _, row in subset.iterrows():
                d, c = float(row["Días"]), float(row["Cantidad"])
                if es_control:
                    cu = float(row["Costo Unitario"]); ct = d * c * cu; vu = float(row["Precio Unitario"]); vt = float(row["Subtotal Item"]); gan = vt - ct; pct = (gan / vt) if vt > 0 else 0
                    vals = [row["Descripción"], d, c, cu, ct, vu, vt, gan, pct]
                    sub_costo += ct; sub_venta += vt; sub_gan += gan
                else:
                    vals = [row["Descripción"], d, c, float(row["Precio Unitario"]), float(row["Subtotal Item"])]
                for idx, v in enumerate(vals, 1):
                    cell = ws.cell(curr_row, idx, v); cell.border = border
                    if es_control and headers[idx-1] == "% GANANCIA":
                        cell.number_format = '0.0%'
                        if isinstance(v, (int, float)):
                            if v >= 0.35: cell.fill, cell.font = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"), Font(color="006100")
                            else: cell.fill, cell.font = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), Font(color="9C0006")
                    elif any(x in headers[idx-1] for x in ["UNIT", "TOTAL", "VENTA", "GANANCIA", "SUBTOTAL", "PRECIO"]):
                        cell.number_format = currency_format
                curr_row += 1
            
            if es_control:
                sub_pct = (sub_gan / sub_venta) if sub_venta > 0 else 0
                vals_sub = ["", "", "", "Subtotales:", sub_costo, "", sub_venta, sub_gan, sub_pct]
                for idx, v in enumerate(vals_sub, 1):
                    cell = ws.cell(curr_row, idx, v)
                    if v != "":
                        cell.font, cell.border = Font(bold=True), border
                        if idx in [5, 7, 8]: cell.number_format = currency_format
                        elif idx == 9:
                            cell.number_format = '0.0%'
                            if isinstance(v, (int, float)):
                                if v >= 0.35: cell.fill, cell.font = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"), Font(bold=True, color="006100")
                                else: cell.fill, cell.font = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), Font(bold=True, color="9C0006")
                total_costo_general += sub_costo; total_venta_general += sub_venta; total_ganancia_general += sub_gan
            curr_row += 1

    # TOTALES EN EL CUADRO DE CONTROL
    if es_control:
        curr_row += 1
        tot_pct = (total_ganancia_general / total_venta_general) if total_venta_general > 0 else 0
        vals_tot = ["", "", "", "TOTALES GENERALES:", total_costo_general, "", total_venta_general, total_ganancia_general, tot_pct]
        
        # Fila de Totales Generales Destacada
        for idx, v in enumerate(vals_tot, 1):
            cell = ws.cell(curr_row, idx, v)
            if v != "":
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid") # Fondo Azul Oscuro
                cell.border = border
                if idx in [5, 7, 8]: cell.number_format = currency_format
                elif idx == 9:
                    cell.number_format = '0.0%'
                    if isinstance(v, (int, float)):
                        if v >= 0.35: cell.font = Font(bold=True, color="00FF00") # Verde Neón
                        else: cell.font = Font(bold=True, color="FF9999") # Rojo Claro
        
        curr_row += 2
        # Resumen Comercial en el Excel de Control
        ws.cell(curr_row, 6, "RESUMEN FINAL CLIENTE").font = Font(bold=True, color="2980b9", size=12)
        curr_row += 1
        resumen_control = [("SUBTOTAL NETO", subtotal_neto), (f"FEE {fee_pct_final}%", fee_total), (f"IVA {iva_porcentaje}%", iva_total), ("TOTAL GENERAL", total_general)]
        for label, val in resumen_control:
            ws.cell(curr_row, 6, label).font = Font(bold=True)
            cv = ws.cell(curr_row, 7, val); cv.number_format = currency_format; cv.font = Font(bold=True)
            curr_row += 1

    # TOTALES EN LA COTIZACIÓN COMERCIAL
    if not es_control:
        resumen = [("SUBTOTAL", subtotal_neto), (f"FEE {fee_pct_final}%", fee_total), (f"IVA {iva_porcentaje}%", iva_total), ("TOTAL GENERAL", total_general)]
        for label, val in resumen:
            ws.cell(curr_row, 4, label).font = Font(bold=True)
            cv = ws.cell(curr_row, 5, val); cv.number_format = currency_format; cv.font = Font(bold=True); curr_row += 1

    # Memoria Oculta para recarga
    ws_raw = wb.create_sheet("ELYON_BACKUP")
    ws_raw.sheet_state = 'hidden'
    datos_generales = {"cotizacion_num": cotizacion_num, "ciudad": ciudad, "fecha_emision": fecha_emision.strftime('%Y-%m-%d'), "validez_dias": validez_dias, "sin_fee": sin_fee, "fee_porcentaje_input": fee_porcentaje_input, "iva_porcentaje": iva_porcentaje, "solicitante_empresa": solicitante_empresa, "solicitante_contacto": solicitante_contacto, "solicitante_email": solicitante_email, "solicitante_telefono": solicitante_telefono, "evento_referencia": evento_referencia, "evento_fecha": evento_fecha.strftime('%Y-%m-%d'), "evento_lugar": evento_lugar, "evento_asistentes": evento_asistentes}
    for r_idx, (k, v) in enumerate(datos_generales.items(), 1):
        ws_raw.cell(row=r_idx, column=1, value=k); ws_raw.cell(row=r_idx, column=2, value=str(v))
    if not df_data.empty:
        for col_idx, col_name in enumerate(df_data.columns, 1): ws_raw.cell(row=20, column=col_idx, value=col_name)
        for r_idx, row_data in enumerate(df_data.values, 21):
            for c_idx, val in enumerate(row_data, 1): ws_raw.cell(row=r_idx, column=c_idx, value=val)
    
    for col in ws.columns:
        max_len = 0; column = col[0].column_letter
        for cell in col:
            try: 
                if len(str(cell.value)) > max_len: max_len = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_len + 5
    output = BytesIO(); wb.save(output); return output.getvalue()

# --- GENERADOR PDF ---
def generar_pdf():
    pdf = PDF(format='letter'); pdf.add_page()
    wm_b, logo_b = obtener_imagen_bytes(URL_WATERMARK_PDF), obtener_imagen_bytes(URL_LOGO_PDF)
    if wm_b:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as t: t.write(wm_b); t_p = t.name
        pdf.image(t_p, x=68, y=100, w=80); os.remove(t_p)
    if logo_b:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as t: t.write(logo_b); t_p = t.name
        pdf.image(t_p, x=10, y=10, w=60); os.remove(t_p)
    
    pdf.set_font("Arial", 'B', 16); pdf.cell(190, 10, "COTIZACIÓN COMERCIAL", ln=True, align="R")
    pdf.set_font("Arial", '', 10); pdf.cell(190, 5, f"Referencia: {cotizacion_num}", ln=True, align="R"); pdf.ln(15)
    
    pdf.set_fill_color(0, 143, 57); pdf.set_text_color(255, 255, 255); pdf.set_font("Arial", 'B', 11)
    pdf.cell(190, 8, " INFORMACIÓN GENERAL", ln=True, fill=True)
    pdf.set_text_color(0, 0, 0); pdf.set_font("Arial", '', 10)
    pdf.cell(100, 8, f" Cliente: {solicitante_empresa}", border='LT'); pdf.cell(90, 8, f" Contacto: {solicitante_contacto}", border='RT', ln=True)
    pdf.cell(190, 8, f" Ciudad: {ciudad}  |  Email: {solicitante_email}", border='LR', ln=True)
    pdf.cell(190, 8, f" Teléfono: {solicitante_telefono}", border='LRB', ln=True); pdf.ln(5)
    
    for cat in st.session_state.cat_order:
        subset = df_validos[df_validos["Categoría"] == cat] if not df_validos.empty else pd.DataFrame()
        if not subset.empty:
            pdf.ln(8); pdf.set_fill_color(230, 230, 230); pdf.set_font("Arial", 'B', 10); pdf.cell(190, 8, f" SECCIÓN: {cat.upper()}", ln=True, fill=True, border=1)
            pdf.set_fill_color(30, 30, 30); pdf.set_text_color(255, 255, 255); pdf.set_font("Arial", 'B', 8)
            pdf.cell(75, 8, "DESCRIPCIÓN", 1, 0, 'L', True); pdf.cell(15, 8, "DIAS", 1, 0, 'C', True); pdf.cell(20, 8, "CANT.", 1, 0, 'C', True); pdf.cell(40, 8, "UNITARIO", 1, 0, 'C', True); pdf.cell(40, 8, "SUBTOTAL", 1, 1, 'C', True)
            pdf.set_text_color(0, 0, 0); pdf.set_font("Arial", '', 8)
            
            for _, r in subset.iterrows():
                desc = str(r["Descripción"])
                lineas_totales = 0
                for parrafo in desc.split('\n'):
                    ancho_p = pdf.get_string_width(parrafo)
                    lineas_totales += max(1, math.ceil(ancho_p / 72.0))
                h_linea = 5 if lineas_totales > 1 else 8
                altura_estimada = max(8, lineas_totales * 5)
                if pdf.get_y() + altura_estimada > 260: pdf.add_page()
                
                x_start, y_start = pdf.get_x(), pdf.get_y()
                pdf.multi_cell(75, h_linea, desc, border=0, align='L')
                altura_real = max(8, pdf.get_y() - y_start)
                pdf.rect(x_start, y_start, 75, altura_real)
                
                pdf.set_xy(x_start + 75, y_start)
                pdf.cell(15, altura_real, str(int(r["Días"])), 1, 0, 'C')
                pdf.cell(20, altura_real, str(int(r["Cantidad"])), 1, 0, 'C')
                pdf.cell(40, altura_real, formato_moneda(r['Precio Unitario']), 1, 0, 'R')
                pdf.cell(40, altura_real, formato_moneda(r['Subtotal Item']), 1, 1, 'R')
                pdf.set_y(y_start + altura_real)
                
            pdf.set_font("Arial", 'B', 9); pdf.cell(150, 8, f"Subtotal {cat}:", 1, 0, 'R'); pdf.cell(40, 8, formato_moneda(redondear_al_mil(subset['Subtotal Item'].sum())), 1, 1, 'R')
            
    pdf.ln(10); pdf.set_font("Arial", 'B', 11); pdf.set_fill_color(230, 230, 230)
    if pdf.get_y() + 40 > 255: pdf.add_page()
    pdf.cell(150, 8, "SUBTOTAL", 1, 0, 'R', fill=True); pdf.cell(40, 8, formato_moneda(subtotal_neto), 1, 1, 'R')
    pdf.cell(150, 8, f"FEE {fee_pct_final}%", 1, 0, 'R', fill=True); pdf.cell(40, 8, formato_moneda(fee_total), 1, 1, 'R')
    pdf.cell(150, 8, f"IVA {iva_porcentaje}%", 1, 0, 'R', fill=True); pdf.cell(40, 8, formato_moneda(iva_total), 1, 1, 'R')
    pdf.set_text_color(255, 255, 255); pdf.set_fill_color(0, 143, 57); pdf.set_font("Arial", 'B', 12)
    pdf.cell(150, 10, "VALOR TOTAL NETO (ANTES DE IVA)", 1, 0, 'R', fill=True); pdf.cell(40, 10, formato_moneda(redondear_al_mil(subtotal_neto + fee_total)), 1, 1, 'R', fill=True)
    pdf.ln(3); pdf.cell(150, 10, "TOTAL GENERAL", 1, 0, 'R', fill=True); pdf.cell(40, 10, formato_moneda(total_general), 1, 1, 'R', fill=True)
    
    # --- TÉRMINOS Y CONDICIONES (COMPLETOS) ---
    pdf.ln(10)
    if pdf.get_y() + 35 > 250: 
        pdf.add_page()
        
    pdf.set_text_color(100, 100, 100)
    pdf.set_font("Arial", 'B', 6)
    pdf.cell(0, 4, "TÉRMINOS Y CONDICIONES GENERALES DEL SERVICIO", ln=True)
    pdf.set_font("Arial", '', 6)
    
    texto_terminos = (
        "1. Alcance del servicio: La cotización incluye únicamente lo descrito en la propuesta. Servicios o requerimientos adicionales serán cotizados y facturados por separado.\n"
        "2. Cambios o ajustes: Cualquier modificación técnica, conceptual, logística o de alcance solicitada por el cliente generará costos adicionales.\n"
        "3. Orden de Compra y ejecución: El proyecto se iniciará únicamente con la recepción de la Orden de Compra (OC) y el anticipo acordado. Cualquier retraso en estos procesos extenderá los tiempos de entrega y libera a ELYON PRODUCCIONES S.A.S. de compromisos de fecha.\n"
        "4. Suspensión o cancelación: En caso de suspensión o cancelación del proyecto, el cliente deberá asumir la totalidad de los costos ejecutados, materiales adquiridos, gastos administrativos y penalidades de proveedores, si aplican.\n"
        "5. Forma de pago y mora: El plazo de pago es de hasta 30 días calendario a partir de la fecha de emisión de la factura, salvo acuerdo escrito. El no pago oportuno generará intereses de mora a la tasa máxima legal vigente en Colombia, más los costos administrativos y de cobranza, hasta la cancelación total de la obligación.\n"
        "6. Retrasos atribuibles al cliente: Retrasos en aprobaciones, entrega de información, artes, documentos o pagos afectarán el cronograma sin responsabilidad para ELYON PRODUCCIONES S.A.S.\n"
        "7. Validez de la cotización: La cotización tiene una vigencia de 15 días calendario. Vencido este plazo, los valores podrán ser ajustados.\n"
        "8. Aceptación y mérito ejecutivo: La aceptación de la cotización por cualquier medio implica la aceptación total de estos términos y condiciones y el reconocimiento expreso de la obligación de pago. La cotización aprobada y la factura correspondiente prestan mérito ejecutivo para efectos de cobro judicial conforme a la ley."
    )
    pdf.multi_cell(0, 3, texto_terminos)
    
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf"); pdf.output(tmp.name); tmp.close()
    with open(tmp.name, "rb") as f: b = f.read()
    os.remove(tmp.name); return b

# --- IV. EXPORTAR ---
st.markdown("<hr>", unsafe_allow_html=True)
st.markdown("### IV. EXPORTAR DOCUMENTOS")
listo = not df_validos.empty
col_b1, col_b2, col_b3 = st.columns(3)
empresa = solicitante_empresa if solicitante_empresa else "Cliente"
pdf_data = generar_pdf() if listo else b""
xlsx_cot = generar_excel_estilizado(df_validos, False) if listo else b""
xlsx_ctrl = generar_excel_estilizado(df_validos, True) if listo else b""
filename = f"Cot_{cotizacion_num}_{empresa}"

col_b1.download_button("📥 COT PDF", pdf_data, f"{filename}.pdf", "application/pdf", disabled=not listo, use_container_width=True)
col_b2.download_button("📊 COT EXCEL", xlsx_cot, f"{filename}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", disabled=not listo, use_container_width=True)
col_b3.download_button("📈 CUADRO CONTROL", xlsx_ctrl, f"Control_{cotizacion_num}_{empresa}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", disabled=not listo, use_container_width=True)
st.markdown("<div style='text-align: center; font-size: 10px; color: #444746; padding-top: 60px;'>by Oscar Buitrago / Elyon Producciones.</div>", unsafe_allow_html=True)

# --- BOTÓN DE GUARDADO RÁPIDO EN BARRA LATERAL ---
st.sidebar.markdown("---")
st.sidebar.markdown("""
<div class="sidebar-card" style="margin-top: 5px;">
    <h3 style='color: #2980b9; margin-bottom: 8px; font-size: 1.1em;'>💾 GUARDAR PROGRESO</h3>
    <p style='color: #bbb; font-size: 0.85em; line-height: 1.4; margin-bottom: 0;'>¿Llevas mucho tiempo editando? Descarga un respaldo rápido para asegurar tus datos contra apagones.</p>
</div>
""", unsafe_allow_html=True)
borrador_xlsx = generar_excel_estilizado(df_validos, False)
st.sidebar.download_button("📥 DESCARGAR BORRADOR RÁPIDO", borrador_xlsx, f"Borrador_Elyon.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
